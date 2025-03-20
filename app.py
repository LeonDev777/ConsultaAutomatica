import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep


planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes.active  # Pega a aba ativa


try:
    planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
    pagina_fechamento = planilha_fechamento.active  # Pega a aba ativa
except FileNotFoundError:
    planilha_fechamento = openpyxl.Workbook()  # Criar nova planilha se não existir
    pagina_fechamento = planilha_fechamento.active
    pagina_fechamento.title = "Sheet1"
    pagina_fechamento.append(["Nome", "Valor", "CPF", "Vencimento", "Status", "Data Pagamento", "Método Pagamento"])


driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(3)


for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    print(f"Consultando CPF: {cpf}")  

    try:
       
        campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
        sleep(1)
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(cpf)
        sleep(1)

       
        botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
        sleep(1)
        botao_pesquisar.click()
        sleep(4)

       
        status_element = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
        status = status_element.text.strip()

        if status == 'em dia':
            data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']").text.split()[3]
            metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']").text.split()[3]
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento, metodo_pagamento])
        else:
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

       
        planilha_fechamento.save('planilha_fechamento.xlsx')

    except Exception as e:
        print(f"Erro ao processar CPF {cpf}: {e}")


input("Pressione Enter para fechar o navegador...")
driver.quit()
